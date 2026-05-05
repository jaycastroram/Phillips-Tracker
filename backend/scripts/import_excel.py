from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))
from app import DB_PATH, IS_POSTGRES, connect, execute, init_db  # noqa: E402


DEFAULT_WORKBOOK = Path(
    r"C:\Users\FSIJonathanCastro-Ra\Downloads\Phillips Project Tracker - Copy for FSI Dev.xlsx"
)

SHEET_MAP = {
    "Ad Hoc": "ad-hoc",
    "Buys": "buys",
    "Completed": "completed",
}

FIELD_BY_COLUMN = {
    1: "date_or_buy",
    2: "current_status",
    3: "visual_reference",
    4: "brand",
    5: "program_name",
    6: "item_name",
    7: "qty",
    8: "important_notes",
    9: "mrl_order_number",
    10: "estimated_ship_date",
    11: "estimated_ihd",
    12: "tracking",
}


def normalize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def import_workbook(workbook_path: Path) -> None:
    init_db()
    workbook = load_workbook(workbook_path, data_only=True)

    with connect() as conn:
        execute(conn, "DELETE FROM tracker_items")

        imported = 0
        for sheet_name, sheet_key in SHEET_MAP.items():
            ws = workbook[sheet_name]
            headers = {
                col: normalize(ws.cell(1, col).value)
                for col in range(13, ws.max_column + 1)
            }

            for row_idx in range(2, ws.max_row + 1):
                row_values = {
                    field: normalize(ws.cell(row_idx, col).value)
                    for col, field in FIELD_BY_COLUMN.items()
                }
                extra = {
                    headers[col] or f"Column {col}": normalize(ws.cell(row_idx, col).value)
                    for col in range(13, ws.max_column + 1)
                    if normalize(ws.cell(row_idx, col).value)
                }

                if not any(row_values.values()) and not extra:
                    continue

                execute(
                    conn,
                    """
                    INSERT INTO tracker_items (
                        sheet, source_row, sort_order, date_or_buy, current_status,
                        visual_reference, brand, program_name, item_name, qty,
                        important_notes, mrl_order_number, estimated_ship_date,
                        estimated_ihd, tracking, extra_json
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sheet_key,
                        row_idx,
                        row_idx - 1,
                        row_values["date_or_buy"],
                        row_values["current_status"],
                        row_values["visual_reference"],
                        row_values["brand"],
                        row_values["program_name"],
                        row_values["item_name"],
                        row_values["qty"],
                        row_values["important_notes"],
                        row_values["mrl_order_number"],
                        row_values["estimated_ship_date"],
                        row_values["estimated_ihd"],
                        row_values["tracking"],
                        json.dumps(extra),
                    ),
                )
                imported += 1

    target = "PostgreSQL DATABASE_URL" if IS_POSTGRES else str(DB_PATH)
    print(f"Imported {imported} tracker rows into {target}")


if __name__ == "__main__":
    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    import_workbook(workbook_arg)
